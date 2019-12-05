import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

# check：X*1636；src:W*9707;dst:R*8094

# 初始化
wbsrc = load_workbook("src.xlsx")
wssrc = wbsrc.active
wbcheck = load_workbook("check.xlsx")
wscheck = wbcheck.active
wbdst = load_workbook("dst.xlsx")
wsdst = wbdst.active
wbnew = load_workbook("new.xlsx")
wsnew = wbnew.active

# 遍历I列
for i in range(2, 3):  # 9708
    a = wssrc.cell(i, 9).value
    if a[0] == "G":
        xulie = 9
    else:
        xulie = 10
    # 检索是否已存在
    dstlen = wsdst.max_row + 1
    nexist1 = True
    for j in range(2, dstlen):
        if a == wsdst.cell(j, xulie).value:
            print(j)
            nexist1 = False
            break
    # 检查是否已有，未有则继续遍历
    if nexist1:
        nexist2 = True
        for k in range(2, 1637):
            if a == wscheck.cell(k, xulie).value:
                print(k)
                for len1 in range(1, 25):
                    wsdst.cell(dstlen, len1).value = wscheck.cell(wscheck.max_row + 1, len1).value
                    nexist2 = False
                break
        if nexist2:
            for len2 in range(1, 25):
                wsnew.cell(dstlen, len2).value = wscheck.cell(wscheck.max_row + 1, len2).value
wbnew.save("newnew.xlsx")
wbdst.save("newdst.xlsx")

'''
# 超长遗留工单整理：1为源数据；2为区域、情况查询表；3为集客。
def sortout_leftorder(url1, url2, url3):
    wb1 = xw.Book(url1)
    wb2 = xw.Book(url2)
    wb3 = xw.Book(url3)

    w1s1 = wb1.sheets[0]


if __name__ == '__main__':
    # 初始化参数
    orifile = "C:/Users/24441/Desktop/"
    storfile = "C:/data/资料/移动/业务类/遗留故障超长工单/"

wb = xw.Book()
wb.save(storfile + "遗留清单" + str(datetime.date.today()) + ".xlsx")
wb.close()
print(datetime.date.today())
'''
